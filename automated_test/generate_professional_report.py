#!/usr/bin/env python3
"""
Professional DAST Report Generator
Converts raw test results into executive-ready Excel with:
- Test ID, Endpoint, Method, Role/Auth Type
- Expected vs Actual Status/Response
- PASS/FAIL indicator with color coding
- Response time analysis
- Detailed findings and remediation steps
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Remediation mapping for common issues
REMEDIATION_GUIDE = {
    'rate_limiting': {
        'issue': 'No rate limit detected - API allows unlimited requests',
        'impact': 'DoS attacks, brute force password attempts, API abuse',
        'fix': 'Implement rate limiting: 100 requests/minute per user, 1000/hour globally',
        'code': 'Use middleware like flask-limiter or implement custom decorator'
    },
    'parameter_pollution': {
        'issue': 'HTTP Parameter Pollution allows duplicate parameters to bypass validation',
        'impact': 'Logic bypass, WAF evasion, security filter circumvention',
        'fix': 'Sanitize and validate all input parameters, use strict parameter parsing',
        'code': 'Reject requests with duplicate parameters or use first/last value consistently'
    },
    'authn_bypass': {
        'issue': 'Protected endpoints accessible without authentication',
        'impact': 'Unauthorized data access, privilege escalation',
        'fix': 'Enforce authentication on all protected endpoints, validate tokens',
        'code': 'Add @auth_required decorator to all protected endpoints'
    },
    'authz_bypass': {
        'issue': 'Role-based access control not properly enforced',
        'impact': 'Privilege escalation, unauthorized access to admin features',
        'fix': 'Implement RBAC checks on all protected operations',
        'code': 'Add role verification: if user.role != "admin": raise PermissionError'
    },
    'idor': {
        'issue': 'Insecure Direct Object References - direct ID manipulation accesses others\' data',
        'impact': 'Data breach, privacy violation, unauthorized information disclosure',
        'fix': 'Verify user ownership before granting access to resources',
        'code': 'Check: if resource.user_id != current_user.id: raise PermissionError'
    },
    'token_tampering': {
        'issue': 'JWT tokens not properly validated, tampered tokens accepted',
        'impact': 'Authentication bypass, session hijacking, identity spoofing',
        'fix': 'Validate JWT signature and expiration on every request',
        'code': 'Use jwt.decode(token, secret, algorithms=["HS256"]) with error handling'
    },
    'injection': {
        'issue': 'SQL/NoSQL injection or command injection vulnerability detected',
        'impact': 'Database compromise, data theft, system takeover',
        'fix': 'Use parameterized queries or ORM, never concatenate user input',
        'code': 'Use db.query(Model).filter(Model.id == user_input) instead of string concat'
    }
}

def load_report_data():
    """Load raw DAST test results."""
    with open('dast_report.json') as f:
        return json.load(f)

def get_status_label(status_code):
    """Convert HTTP status code to label."""
    if status_code is None:
        return 'N/A'
    if status_code == 0:
        return 'Connection Error'
    return f'{status_code}'

def determine_pass_fail(test):
    """Determine if test passed or failed."""
    if test.get('finding') == True:
        return 'FAIL'
    return 'PASS'

def get_remediation(category):
    """Get remediation details for a finding."""
    return REMEDIATION_GUIDE.get(category, {})

def create_summary_sheet(ws_summary, data):
    """Create summary statistics sheet."""
    ws_summary.title = "Summary"
    
    # Title and metadata
    ws_summary['A1'] = "SmartStudy DAST Report"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Statistics
    total = len(data)
    findings = len([d for d in data if d.get('finding') == True])
    passed = total - findings
    
    ws_summary['A4'] = "Statistics"
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    ws_summary['A5'] = "Total Tests"
    ws_summary['B5'] = total
    ws_summary['A6'] = "Tests Passed"
    ws_summary['B6'] = passed
    ws_summary['A6'].font = Font(color="008000")  # Green
    
    ws_summary['A7'] = "Tests Failed (Findings)"
    ws_summary['B7'] = findings
    if findings > 0:
        ws_summary['A7'].font = Font(color="FF0000")  # Red
    
    ws_summary['A8'] = "Pass Rate"
    ws_summary['B8'] = f"{(passed/total)*100:.1f}%"
    
    # By Category
    ws_summary['A10'] = "Findings by Category"
    ws_summary['A10'].font = Font(bold=True, size=12)
    
    categories = {}
    for d in data:
        if d.get('finding') == True:
            cat = d['test_category']
            categories[cat] = categories.get(cat, 0) + 1
    
    row = 11
    for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
        ws_summary[f'A{row}'] = cat
        ws_summary[f'B{row}'] = count
        row += 1
    
    # Set column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15

def create_detailed_sheet(ws_detail, data):
    """Create detailed test results sheet."""
    ws_detail.title = "Test Results"
    
    # Headers
    headers = [
        'Test ID', 'Endpoint', 'Method', 'Role/Auth', 
        'Expected Response', 'Actual Response', 'Status',
        'Response Time (ms)', 'Finding', 'Severity', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, test in enumerate(data, 2):
        test_id = idx - 1
        status = determine_pass_fail(test)
        
        row_data = [
            test_id,
            test['endpoint'],
            test['method'],
            test['role'] if test['role'] else 'None',
            test.get('expected_status', 'N/A'),
            get_status_label(test.get('status')),
            status,
            test['response_time_ms'],
            'Yes' if test.get('finding') == True else 'No',
            test['severity'].upper(),
            test.get('note', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_detail.cell(row=idx, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Color code by status
            if col == 7:  # Status column
                if status == 'PASS':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
                cell.alignment = Alignment(horizontal="center")
    
    # Set column widths
    col_widths = [10, 25, 8, 15, 20, 20, 10, 15, 10, 12, 30]
    for idx, width in enumerate(col_widths, 1):
        ws_detail.column_dimensions[chr(64 + idx)].width = width
    
    # Freeze header row
    ws_detail.freeze_panes = "A2"

def create_findings_sheet(ws_findings, data):
    """Create detailed findings and remediation sheet."""
    ws_findings.title = "Findings & Remediation"
    
    # Header
    ws_findings['A1'] = "Security Findings Report"
    ws_findings['A1'].font = Font(size=14, bold=True)
    ws_findings['A2'] = "Detailed findings with remediation guidance"
    ws_findings['A2'].font = Font(italic=True)
    
    findings = [d for d in data if d.get('finding') == True]
    
    if not findings:
        ws_findings['A4'] = "✓ No security findings detected"
        ws_findings['A4'].font = Font(color="008000", size=12, bold=True)
    else:
        row = 4
        for idx, finding in enumerate(findings, 1):
            # Finding header
            ws_findings[f'A{row}'] = f"Finding #{idx}"
            ws_findings[f'A{row}'].font = Font(bold=True, size=11, color="FF0000")
            row += 1
            
            ws_findings[f'A{row}'] = f"Endpoint: {finding['endpoint']}"
            row += 1
            
            ws_findings[f'A{row}'] = f"Category: {finding['test_category']}"
            row += 1
            
            ws_findings[f'A{row}'] = f"Severity: {finding['severity'].upper()}"
            ws_findings[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws_findings[f'A{row}'] = f"Details: {finding.get('note', 'N/A')}"
            ws_findings[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
            
            # Remediation
            remediation = get_remediation(finding['test_category'])
            if remediation:
                ws_findings[f'A{row}'] = "Remediation Steps:"
                ws_findings[f'A{row}'].font = Font(bold=True, underline="single")
                row += 1
                
                ws_findings[f'A{row}'] = f"Issue: {remediation.get('issue', 'N/A')}"
                ws_findings[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1
                
                ws_findings[f'A{row}'] = f"Impact: {remediation.get('impact', 'N/A')}"
                ws_findings[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1
                
                ws_findings[f'A{row}'] = f"Fix: {remediation.get('fix', 'N/A')}"
                ws_findings[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1
                
                ws_findings[f'A{row}'] = f"Code Example: {remediation.get('code', 'N/A')}"
                ws_findings[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws_findings[f'A{row}'].font = Font(italic=True, color="0070C0")
                row += 2
        
        ws_findings.column_dimensions['A'].width = 100

def main():
    print("🔍 Loading DAST test results...")
    data = load_report_data()
    
    print("📊 Creating professional Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    create_summary_sheet(wb.create_sheet(), data)
    create_detailed_sheet(wb.create_sheet(), data)
    create_findings_sheet(wb.create_sheet(), data)
    
    # Save workbook
    output_file = 'dast_report_professional.xlsx'
    wb.save(output_file)
    
    print(f"✓ Professional report saved: {output_file}")
    print(f"\n📈 Report Summary:")
    print(f"   Total Tests: {len(data)}")
    findings = len([d for d in data if d.get('finding') == True])
    print(f"   Findings: {findings}")
    print(f"   Pass Rate: {((len(data)-findings)/len(data)*100):.1f}%")

if __name__ == '__main__':
    main()
