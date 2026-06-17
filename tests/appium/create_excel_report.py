import re, ast, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def load_test_cases(file_path: Path):
    text = file_path.read_text(encoding='utf-8')
    # Extract the list assigned to TEST_CASES
    match = re.search(r"TEST_CASES\s*=\s*(\[.*?\])", text, re.DOTALL)
    if not match:
        raise ValueError('Could not find TEST_CASES definition')
    list_str = match.group(1)
    # Safely evaluate the list of dicts
    test_cases = ast.literal_eval(list_str)
    return test_cases

def generate_excel(test_cases, output_path: Path):
    wb = Workbook()
    # Dashboard summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Dashboard'
    total = len(test_cases)
    passed = sum(1 for tc in test_cases if tc.get('status') == 'PASS')
    failed = total - passed
    ws_summary.append(['Total Test Cases', total])
    ws_summary.append(['Passed', passed])
    ws_summary.append(['Failed', failed])
    for cell in ws_summary["A1":"B3"]:
        for c in cell:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
    # Pass/Fail sheet
    ws_pf = wb.create_sheet(title='Pass_Fail')
    ws_pf.append(['ID', 'Name', 'Status', 'Error'])
    for tc in test_cases:
        ws_pf.append([
            tc.get('id'),
            tc.get('name'),
            tc.get('status'),
            tc.get('error', '')
        ])
    # Only Passed sheet
    ws_pass = wb.create_sheet(title='Only_Passed')
    ws_pass.append(['ID', 'Name'])
    for tc in test_cases:
        if tc.get('status') == 'PASS':
            ws_pass.append([tc.get('id'), tc.get('name')])
    wb.save(output_path)

if __name__ == '__main__':
    repo_root = Path(__file__).resolve().parents[2]  # project root
    suite_file = repo_root / 'tests' / 'appium' / 'generate_test_suite_report.py'
    output_file = repo_root / 'tests' / 'appium' / 'E2E_Test_Report.xlsx'
    cases = load_test_cases(suite_file)
    generate_excel(cases, output_file)
    print(f'Excel report generated at {output_file}')
