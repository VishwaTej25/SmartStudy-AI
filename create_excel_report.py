import os
import importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Load the generate_test_suite_report module to get TEST_CASES
def load_test_cases():
    report_path = os.path.join(os.path.dirname(__file__), 'tests', 'appium', 'generate_test_suite_report.py')
    spec = importlib.util.spec_from_file_location('report', report_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return getattr(module, 'TEST_CASES', [])

TEST_CASES = load_test_cases()

# Create workbook
wb = Workbook()

# Dashboard Summary sheet
summary_ws = wb.active
summary_ws.title = 'Dashboard Summary'
summary_ws['A1'] = 'SmartStudy E2E Test Report'
summary_ws['A2'] = 'Generated on'
summary_ws['B2'] = os.path.abspath(__file__)
summary_ws['A4'] = 'Total Test Cases'
summary_ws['B4'] = len(TEST_CASES)
passed = sum(1 for tc in TEST_CASES if tc.get('status') == 'PASS')
failed = sum(1 for tc in TEST_CASES if tc.get('status') == 'FAIL')
summary_ws['A5'] = 'Passed'
summary_ws['B5'] = passed
summary_ws['A6'] = 'Failed'
summary_ws['B6'] = failed
# Formatting (optional)
for cell in ['A1','A2','A4','A5','A6']:
    summary_ws[cell].font = Font(bold=True)

# Pass & Fail sheet
pf_ws = wb.create_sheet('Pass & Fail Testcases')
headers = ['ID','Category','Name','Expected','Status','Time','Error']
pf_ws.append(headers)
for tc in TEST_CASES:
    pf_ws.append([
        tc.get('id'),
        tc.get('category'),
        tc.get('name'),
        tc.get('expected'),
        tc.get('status'),
        tc.get('time'),
        tc.get('error','')
    ])

# Only Passed sheet
passed_ws = wb.create_sheet('Only Passed')
passed_ws.append(headers)
for tc in TEST_CASES:
    if tc.get('status') == 'PASS':
        passed_ws.append([
            tc.get('id'),
            tc.get('category'),
            tc.get('name'),
            tc.get('expected'),
            tc.get('status'),
            tc.get('time'),
            tc.get('error','')
        ])

# Save workbook
output_path = os.path.join(os.path.dirname(__file__), 'SmartStudy_E2E_Test_Report_Final.xlsx')
wb.save(output_path)
print(f'Excel report generated at {output_path}')
